"""Build / extend the offline sample dataset from live Upstox data.

The bundled workbook (`sample_data/nifty_1y_1min.xlsx`) has two sheets:

    Spot_1min:        Date, Time, Open, High, Low, Close, Volume
    ATM_Options_1min: Date, Time, Timestamp, Strike, Type, Expiry,
                      Open, High, Low, Close, Volume, OI

The original snapshot carried a single ATM strike per day. This script *extends*
it: for each trading day in a window it adds the neighbouring strikes
(ATM +/- N steps, both CE and PE) pulled from Upstox's expired-option candles, so
that same-side spreads (verticals, iron condors) become testable offline. Every
other row in the workbook is preserved untouched, and days outside the window
stay ATM-only (the reader falls back to the nearest available strike).

Usage:
    export UPSTOX_TOKEN=...            # a live token; expired-option candles need auth
    python scripts/build_sample_data.py --start 2026-06-01 --end 2026-06-30

    # wider chain, explicit output
    python scripts/build_sample_data.py --start 2026-06-01 --end 2026-06-30 \
        --offsets " -3,-2,-1,1,2,3" --output sample_data/nifty_1y_1min.xlsx
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date
from pathlib import Path

import openpyxl
import polars as pl

# Make `bhav` importable when run as a plain script.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from bhav.data.excel_source import DEFAULT_EXCEL_PATH  # noqa: E402
from bhav.data.underlyings import default_atm_step  # noqa: E402
from bhav.data.upstox_client import OptionContract, UpstoxClient  # noqa: E402

OPT_SHEET = "ATM_Options_1min"
SPOT_SHEET = "Spot_1min"
OPT_HEADERS = [
    "Date", "Time", "Timestamp", "Strike", "Type", "Expiry",
    "Open", "High", "Low", "Close", "Volume", "OI",
]


def _parse_offsets(raw: str) -> list[int]:
    out = []
    for tok in raw.replace(" ", "").split(","):
        if not tok:
            continue
        n = int(tok)
        if n != 0:  # offset 0 (ATM) already lives in the workbook
            out.append(n)
    return sorted(set(out))


def _existing_atm(opt_sheet: pl.DataFrame) -> dict[date, tuple[int, date]]:
    """Per day -> (atm_strike, expiry) taken from the ATM rows already present."""
    rows = (
        opt_sheet.select("Date", "Strike", "Expiry")
        .unique()
        .to_dicts()
    )
    out: dict[date, tuple[int, date]] = {}
    for r in rows:
        d = date.fromisoformat(str(r["Date"]))
        out.setdefault(d, (int(r["Strike"]), date.fromisoformat(str(r["Expiry"]))))
    return out


def _chain_for(client: UpstoxClient, underlying: str, expiry: date,
               cache: dict) -> dict[tuple[int, str], OptionContract]:
    if expiry not in cache:
        contracts = client.get_expired_contracts(underlying, expiry)
        cache[expiry] = {(c.strike, c.option_type): c for c in contracts}
    return cache[expiry]


def _candle_rows(candles: list[list], strike: int, otype: str, expiry: date) -> list[dict]:
    rows = []
    for c in candles:
        ts = c[0]  # "2026-06-01T09:15:00+05:30"
        rows.append({
            "Date": ts[:10],
            "Time": ts[11:19],
            "Timestamp": ts,
            "Strike": strike,
            "Type": otype,
            "Expiry": expiry.isoformat(),
            "Open": float(c[1]), "High": float(c[2]),
            "Low": float(c[3]), "Close": float(c[4]),
            "Volume": int(c[5]), "OI": int(c[6]) if len(c) > 6 else 0,
        })
    return rows


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--start", required=True, help="YYYY-MM-DD (inclusive)")
    ap.add_argument("--end", required=True, help="YYYY-MM-DD (inclusive)")
    ap.add_argument("--offsets", default="-2,-1,1,2", help="strike steps to add around ATM")
    ap.add_argument("--underlying", default="NSE_INDEX|Nifty 50")
    ap.add_argument("--input", default=str(DEFAULT_EXCEL_PATH))
    ap.add_argument("--output", default=None, help="defaults to --input (in place)")
    ap.add_argument("--interval", default="1minute")
    ap.add_argument("--token", default=os.environ.get("UPSTOX_TOKEN"))
    args = ap.parse_args()

    if not args.token:
        ap.error("a live Upstox token is required (set UPSTOX_TOKEN or pass --token)")

    start, end = date.fromisoformat(args.start), date.fromisoformat(args.end)
    offsets = _parse_offsets(args.offsets)
    step = default_atm_step(args.underlying)
    in_path = Path(args.input)
    out_path = Path(args.output) if args.output else in_path

    print(f"Reading {in_path} ...")
    existing = pl.read_excel(in_path, sheet_name=OPT_SHEET)
    spot = pl.read_excel(in_path, sheet_name=SPOT_SHEET)
    atm_by_day = _existing_atm(existing)

    days = sorted(d for d in atm_by_day if start <= d <= end)
    print(f"{len(days)} trading days in window; adding offsets {offsets} (step {step}).")

    have = set(
        (date.fromisoformat(str(r["Date"])), int(r["Strike"]), r["Type"])
        for r in existing.select("Date", "Strike", "Type").unique().to_dicts()
    )

    new_rows: list[dict] = []
    chain_cache: dict = {}
    with UpstoxClient(args.token) as client:
        for i, d in enumerate(days, 1):
            atm, expiry = atm_by_day[d]
            chain = _chain_for(client, args.underlying, expiry, chain_cache)
            for off in offsets:
                strike = atm + off * step
                for otype in ("CE", "PE"):
                    if (d, strike, otype) in have:
                        continue
                    contract = chain.get((strike, otype))
                    if contract is None:
                        continue
                    candles = client.get_expired_option_candles(
                        contract.instrument_key, args.interval, d, d
                    )
                    new_rows.extend(_candle_rows(candles, strike, otype, expiry))
            if i % 5 == 0 or i == len(days):
                print(f"  {i}/{len(days)} days, {len(new_rows)} new rows so far")

    if not new_rows:
        print("No new rows fetched (window already complete or no data). Nothing written.")
        return 0

    combined = (
        pl.concat([existing.select(OPT_HEADERS), pl.DataFrame(new_rows).select(OPT_HEADERS)])
        .unique(subset=["Date", "Timestamp", "Strike", "Type"], keep="last")
        .sort(["Date", "Type", "Strike", "Timestamp"])
    )
    print(f"Options rows: {existing.height} -> {combined.height}. Writing {out_path} ...")
    _write_workbook(out_path, spot, combined)
    print("Done.")
    return 0


def _write_workbook(path: Path, spot: pl.DataFrame, options: pl.DataFrame) -> None:
    wb = openpyxl.Workbook()
    ws_spot = wb.active
    ws_spot.title = SPOT_SHEET
    ws_spot.append(list(spot.columns))
    for row in spot.iter_rows():
        ws_spot.append(list(row))
    ws_opt = wb.create_sheet(OPT_SHEET)
    ws_opt.append(list(options.columns))
    for row in options.iter_rows():
        ws_opt.append(list(row))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    raise SystemExit(main())
